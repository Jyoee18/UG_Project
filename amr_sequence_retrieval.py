# -*- coding: utf-8 -*-
"""
@author: jyothishree

Script Name: amr_sequence_retrieval.py

Version: 3.00

Description:
This script processes sequences in a FASTA file by extracting specific subsequences(AMR sequences) based on start and end positions provided in an accompanying Excel file. It outputs the extracted sequences, their lengths, and any associated resistance information into a new Excel file. 

The script requires two input files:
1. A FASTA file containing sequences to process.
2. An Excel (.xlsx) file with columns indicating sequence identifiers, start and end positions, and resistance values.

Output:
    A new Excel file is generated with extracted AMR sequence information, including start and end positions, resistance data, and sequence lengths.

Usage:
    python amr_sequence_retrieval.py <input_fasta> <start_end_xlsx>

Example:
    python amr_sequence_retrieval.py sequences.fasta start_end_data.xlsx

Error Handling:
    1. FileNotFoundError: Displays an error message if the input FASTA or Excel file is not found.
    2. ValueError: Shows an error if data values (start, end) cannot be converted to integers.
    3. KeyError: If the required column is missing from the Excel file, a warning is displayed.
    4. General Exception: Any unexpected errors in loading data or processing are displayed to the user.
"""
# Importing necessary modules
import os  
import sys  
from Bio import SeqIO  
from openpyxl import load_workbook, Workbook  
import time  

def get_cut_sequence(sequence, start, end):
    """
    Extracts a subsequence from a given sequence based on start and end positions.

    Parameters:
        sequence (str): The full DNA or RNA sequence.
        start (int): The start position of the subsequence.
        end (int): The end position of the subsequence.

    Returns:
        str: The extracted subsequence.
    """
    return sequence[start - 1:end]  # Returns the subsequence from start to end position

def process_xlsx(xlsx_file):
    """
    Reads start, end, and resistance data from an Excel file.

    Parameters:
        xlsx_file (str): Path to the Excel (.xlsx) file.

    Returns:
        dict: A dictionary with sequence identifiers as keys and a list of tuples
              (start, end, resistance) as values, or None if there's an error.
    """
    data = {}  # Initializes an empty dictionary to store data
    try:
        workbook = load_workbook(xlsx_file)  # Opens the Excel workbook
        sheet = workbook.active  # Gets the active sheet in the workbook
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Iterates over rows, starting from the second row
            sequence_identifier, start, end, resistance = row[1], row[2], row[3], row[12]  # Reads columns for each row
            if sequence_identifier and start and end and resistance:  # Checks if all required values are present
                if sequence_identifier in data:
                    data[sequence_identifier].append((int(start), int(end), resistance))  # Adds to existing entry
                else:
                    data[sequence_identifier] = [(int(start), int(end), resistance)]  # Creates a new entry
    except Exception as e:
        print(f"Error loading data from Excel file: {e}")  # Prints error message if there's an exception
        return None  # Returns None if an error occurs
    return data  # Returns the data dictionary

def process_fasta(input_fasta, start_end_xlsx):
    """
    Processes the input FASTA file and extracts sequences based on start and end data 
    from the provided Excel file, then saves the results to an output Excel file in 
    the same directory as the input FASTA file.

    Parameters:
        input_fasta (str): Path to the input FASTA file.
        start_end_xlsx (str): Path to the Excel file containing start, end, and resistance data.
    """
    try:
        # Get the directory of the input FASTA file
        input_dir = os.path.dirname(input_fasta)  # Gets directory path of the input FASTA file
        input_filename = os.path.basename(input_fasta)  # Gets the file name of the input FASTA
        input_identifier = os.path.splitext(input_filename)[0].split('-')[-1].split('_')[-1].strip('[]')  # Extracts an identifier for output file naming
        
        # Define output path in the same directory as the input files
        output_xlsx = f"{input_identifier}_AMR.xlsx"  # Constructs the output file name
        output_path = os.path.join(input_dir, output_xlsx)  # Joins directory with output file name

        # Load start and end positions with resistance data from the Excel file
        start_end_dict = process_xlsx(start_end_xlsx)  # Calls function to get data from Excel file
        if start_end_dict is None:  # Checks if loading data was unsuccessful
            print("Error: Failed to load data from Excel file.")
            return

        # Create a new Excel workbook for output
        workbook = Workbook()  # Creates a new workbook
        sheet = workbook.active  # Gets the active sheet in the new workbook
        sheet.append(['Sl. No', 'Input Sequence Identifier', 'Input Sequence', 'Length of Input Sequence',
                      'Start', 'End', 'Sequence from Start and End', 'Length of Cut Sequence', 'Resistance'])  # Adds header row

        # Read and process each sequence in the FASTA file
        with open(input_fasta, "r") as fasta_file:
            records = list(SeqIO.parse(fasta_file, "fasta"))  # Parses the FASTA file and converts to a list of records
            for idx, record in enumerate(records, start=1):  # Loops through each record, starting index from 1
                sequence_identifier = record.id  # Gets the sequence identifier from the record
                sequence = str(record.seq)  # Gets the sequence as a string
                sequence_length = len(sequence)  # Gets the length of the sequence

                # Check if the sequence identifier is in the start-end dictionary
                if sequence_identifier in start_end_dict:
                    for start, end, resistance in start_end_dict[sequence_identifier]:  # Loops through data for each identifier
                        cut_sequence = get_cut_sequence(sequence, start, end)  # Extracts the subsequence
                        sheet.append([idx, f"{sequence_identifier}_{input_identifier}", sequence, sequence_length,
                                      start, end, cut_sequence, len(cut_sequence), resistance])  # Appends row to the output sheet
                else:
                    print(f"Warning: No start and end values found for {sequence_identifier}. Skipping.")  # Warning for missing data

        workbook.save(output_path)  # Saves the output workbook to the specified path
        print(f"Output XLSX file '{output_path}' has been generated successfully.")  # Prints success message

    except FileNotFoundError as fnf_error:
        print(f"File not found: {fnf_error}")  # Error message if file not found
    except Exception as e:
        print(f"An error occurred while processing the FASTA file: {e}")  # Generic error message

def main():
    """
    Main function to handle command-line arguments and call the necessary functions.
    Expects two command-line arguments: the path to a FASTA file and the path to an Excel file.
    """
    if len(sys.argv) != 3:  # Checks if exactly 3 arguments are provided
        print("Usage: python script_name.py <input_fasta> <start_end_xlsx>")  # Prints usage instructions
        sys.exit(1)  # Exits if incorrect arguments

    input_fasta = sys.argv[1]  # Gets the path to the input FASTA file from arguments
    start_end_xlsx = sys.argv[2]  # Gets the path to the Excel file from arguments

    if not os.path.isfile(input_fasta):  # Checks if FASTA file exists
        print(f"Error: Input FASTA file '{input_fasta}' does not exist.")  # Prints error if file not found
        sys.exit(1)  # Exits the script
    if not os.path.isfile(start_end_xlsx):  # Checks if Excel file exists
        print(f"Error: Excel file '{start_end_xlsx}' does not exist.")  # Prints error if file not found
        sys.exit(1)  # Exits the script

    start_time = time.time()  # Records start time of the process
    process_fasta(input_fasta, start_end_xlsx)  # Calls function to process the FASTA file
    elapsed_time = round(time.time() - start_time, 2)  # Calculates elapsed time
    print(f"Sequence Retrieval Complete. Time taken: {elapsed_time} seconds")  # Prints completion message

if __name__ == "__main__":
    main()  # Executes main function if script is run directly

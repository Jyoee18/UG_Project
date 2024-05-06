import os
from tkinter import Tk, Label, Button, Entry, filedialog
from tkinter.ttk import Progressbar
from Bio import SeqIO
from openpyxl import load_workbook, Workbook
import time

def get_cut_sequence(sequence, start, end):
    cut_sequence = sequence[start-1:end]
    return cut_sequence

def process_xlsx(xlsx_file):
    data = {}
    try:
        workbook = load_workbook(xlsx_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            sequence_identifier, start, end, resistance = row[1], row[2], row[3], row[12]  # Resistance is in the 13th column
            if sequence_identifier in data:
                data[sequence_identifier].append((int(start), int(end), resistance))
            else:
                data[sequence_identifier] = [(int(start), int(end), resistance)]

    except Exception as e:
        print(f"Error loading data from XLSX file: {e}")
        return None

    return data

def process_fasta(input_fasta, start_end_xlsx):
    input_filename = os.path.basename(input_fasta)
    input_identifier = os.path.splitext(input_filename)[0].split('-')[-1].split('_')[-1].strip('[]')  # Extracting identifier from the filename
    output_xlsx = f"{input_identifier}_AMR.xlsx"
    output_directory = os.path.join(os.path.expanduser("~"), "Downloads")
    output_path = os.path.join(output_directory, output_xlsx)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Sl. No', 'Input Sequence Identifier', 'Input Sequence', 'Length of Input Sequence', 'Start', 'End', 'Sequence from Start and End', 'Length of Cut Sequence', 'Resistance'])  

    start_end_dict = process_xlsx(start_end_xlsx)
    if start_end_dict is None:
        output_status_label.config(text="Error: Failed to load data from Excel file.")
        return

    with open(input_fasta, "r") as fasta_file:
        records = list(SeqIO.parse(fasta_file, "fasta"))

        for idx, record in enumerate(records, start=1):
            sequence_identifier = record.id
            sequence = str(record.seq)
            sequence_length = len(sequence)

            if sequence_identifier in start_end_dict:
                for start, end, resistance in start_end_dict[sequence_identifier]:
                    cut_sequence = get_cut_sequence(sequence, start, end)
                    sheet.append([idx, f"{sequence_identifier}_{input_identifier}", sequence, sequence_length, start, end, cut_sequence, len(cut_sequence), resistance])  # Include resistance in the appended row
            else:
                print(f"Warning: No start and end values found for {sequence_identifier}. Skipping.")

    workbook.save(output_path)
    output_status_label.config(text=f"Output XLSX file '{output_path}' has been generated successfully.")

def browse_input_file():
    input_fasta = filedialog.askopenfilename(title="Select Input FASTA File")
    input_file_entry.delete(0, "end")
    input_file_entry.insert(0, input_fasta)

def browse_xlsx_file():
    xlsx_file = filedialog.askopenfilename(title="Select Excel File")
    xlsx_file_entry.delete(0, "end")
    xlsx_file_entry.insert(0, xlsx_file)

def process_files():
    process_button.config(text="Retrieving sequences, please wait...", state="disabled")
    start_time = time.time()
    input_fasta = input_file_entry.get()
    start_end_xlsx = xlsx_file_entry.get()

    def process_and_update():
        process_fasta(input_fasta, start_end_xlsx)
        end_time = time.time()
        elapsed_time = round(end_time - start_time, 2)
        process_button.config(text="Retrieve AMR Sequences", state="normal")
        output_status_label.config(text=f"Sequence Retrieval Complete. Time taken: {elapsed_time} seconds")

    root.after(100, process_and_update)
    
# GUI
root = Tk()
root.title("AMR Sequence Retrieval Tool")

# Function to center the window
def center_window(window):
    window.update_idletasks()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    window_width = window.winfo_reqwidth()  
    window_height = window.winfo_reqheight()  
    x_coordinate = (screen_width - window_width) // 2
    y_coordinate = (screen_height - window_height) // 2
    window.geometry("+{}+{}".format(x_coordinate, y_coordinate))

# Window dimensions
window_width = 500
window_height = 250  # Adjusted height to accommodate the new button

# Geometry
root.geometry(f"{window_width}x{window_height}")

# Center the window
center_window(root)

input_file_label = Label(root, text="Input FASTA File:")
input_file_label.grid(row=0, column=0, sticky="e")

input_file_entry = Entry(root, width=50)
input_file_entry.grid(row=0, column=1, padx=10, pady=5, sticky="w")

input_file_button = Button(root, text="Browse", command=browse_input_file)
input_file_button.grid(row=0, column=2)

xlsx_file_label = Label(root, text="Excel File:")
xlsx_file_label.grid(row=1, column=0, sticky="e")

xlsx_file_entry = Entry(root, width=50)
xlsx_file_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

xlsx_file_button = Button(root, text="Browse", command=browse_xlsx_file)
xlsx_file_button.grid(row=1, column=2)

process_button = Button(root, text="Retrieve AMR Sequences", command=process_files)
process_button.grid(row=2, column=1, pady=10)

output_status_label = Label(root, text="")
output_status_label.grid(row=3, columnspan=3)

disabled_button = Button(root, text="Developed in the Dept. of Biotechnology, SVCE", state="disabled")
disabled_button.grid(row=4, columnspan=3, pady=10)

root.mainloop()
